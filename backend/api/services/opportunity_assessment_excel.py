"""
Build / parse Opportunity Assessment Excel templates.

Template source: templates/opportunity_assessment/
Cascade L1→L2→L3→L4 uses Excel named ranges + INDIRECT data validation.
"""

from __future__ import annotations

import re
from collections import defaultdict
from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from api.models import FpoMapping

TEMPLATE_RELATIVE = Path("templates") / "opportunity_assessment" / (
    "Detailed Scoping Document (task-level Opportunity Assessment).xlsx"
)
DATA_START_ROW = 3
MAX_DATA_ROWS = 200
HEADER_ROW = 2
COL_MIGRATION_REQUEST_ID = 1

# Characters replaced both in Python named-range keys and in Excel SUBSTITUTE formulas
SANITIZE_CHARS = [" ", ".", "-", "/", "(", ")", ",", "'", '"', "+", "%", "&", ":", ";"]

FIELD_BY_HEADER = {
    "migration_request_id": "migration_request_id",
    "owner": "owner",
    "product": "product",
    "location": "location",
    "l1": "l1",
    "l2": "l2",
    "l3": "l3",
    "l4": "l4",
    "task name": "task_name",
    "task description": "task_description",
    "task found in the corresponding service catalogue (y/n)": "task_found_in_service_catalog",
    "task found in the corresponding service catalog?": "task_found_in_service_catalog",
    "migratable to gsc as per service catalogue (y/n)": "migratable_to_gsc",
    "migratable to gsc as per service catalog?": "migratable_to_gsc",
    "upstream (tasks, events, input)": "upstream",
    "downstream (task, events, output)": "downstream",
    "risks related to the process/task": "risks_related",
    "complexity (low, medium, high)": "complexity",
    "sop/iop exists?": "sop_iop_exists",
    "training time needed": "training_time_needed",
    "recommended handoff duration": "recommended_handoff_duration",
    "task frequency": "task_frequency",
    "unit of measure": "unit_of_measure",
    "volume": "volume_monthly",
    "volume monthly": "volume_monthly",
    "volume_monthly": "volume_monthly",
    "task time per unit": "task_time_per_unit_min",
    "task time per unit (min)": "task_time_per_unit_min",
    "task_time_per_unit": "task_time_per_unit_min",
    "task_time_per_unit_min": "task_time_per_unit_min",
    "fte calculation": "fte_calculation",
}


def template_source_path() -> Path:
    return Path(settings.BASE_DIR).parent / TEMPLATE_RELATIVE


def sanitize_token(value: str) -> str:
    """Match Excel SUBSTITUTE chain used in INDIRECT formulas (no underscore collapsing)."""
    text = (value or "").strip()
    for ch in SANITIZE_CHARS:
        text = text.replace(ch, "_")
    text = re.sub(r"[^A-Za-z0-9_]", "_", text)
    return text or "Empty"


def named_l2(l1: str) -> str:
    return f"L2_{sanitize_token(l1)}"[:240]


def named_l3(l1: str, l2: str) -> str:
    return f"L3_{sanitize_token(l1)}_{sanitize_token(l2)}"[:240]


def named_l4(l1: str, l2: str, l3: str) -> str:
    return f"L4_{sanitize_token(l1)}_{sanitize_token(l2)}_{sanitize_token(l3)}"[:240]


def excel_sanitize_formula(cell_ref: str) -> str:
    expr = cell_ref
    for ch in SANITIZE_CHARS:
        safe = ch.replace('"', '""')
        expr = f'SUBSTITUTE({expr},"{safe}","_")'
    return expr


def _load_cascade_maps():
    l1_list: list[str] = []
    l2_by_l1: dict[str, list[str]] = defaultdict(list)
    l3_by_l1_l2: dict[tuple[str, str], list[str]] = defaultdict(list)
    l4_by_l1_l2_l3: dict[tuple[str, str, str], list[str]] = defaultdict(list)
    seen_l1 = set()

    for row in FpoMapping.objects.all().order_by("id").values("l1", "l2", "l3", "l4"):
        l1 = (row["l1"] or "").strip()
        l2 = (row["l2"] or "").strip()
        l3 = (row["l3"] or "").strip()
        l4 = (row["l4"] or "").strip()
        if not l1:
            continue
        if l1 not in seen_l1:
            seen_l1.add(l1)
            l1_list.append(l1)
        if l2 and l2 not in l2_by_l1[l1]:
            l2_by_l1[l1].append(l2)
        if l2 and l3 and l3 not in l3_by_l1_l2[(l1, l2)]:
            l3_by_l1_l2[(l1, l2)].append(l3)
        if l2 and l3 and l4 and l4 not in l4_by_l1_l2_l3[(l1, l2, l3)]:
            l4_by_l1_l2_l3[(l1, l2, l3)].append(l4)

    return l1_list, l2_by_l1, l3_by_l1_l2, l4_by_l1_l2_l3


def _write_list_column(ws, col_idx: int, header: str, values: list[str]) -> str:
    ws.cell(1, col_idx, header)
    for i, value in enumerate(values, start=2):
        ws.cell(i, col_idx, value)
    end_row = max(2, len(values) + 1)
    col_letter = get_column_letter(col_idx)
    return f"'{ws.title}'!${col_letter}$2:${col_letter}${end_row}"


def _add_defined_name(wb, name: str, attr_text: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name=name, attr_text=attr_text))


def build_opportunity_template(migration_request_id: str) -> BytesIO:
    source = template_source_path()
    if not source.exists():
        raise FileNotFoundError(f"Template not found: {source}")

    wb = load_workbook(source)
    ws = wb.active
    ws.title = "Scope"

    end_row = DATA_START_ROW + MAX_DATA_ROWS - 1
    # Clear sample data, formulas, and hyperlinks in the data area
    for row in range(DATA_START_ROW, max(ws.max_row, end_row) + 1):
        for col in range(1, 23):
            cell = ws.cell(row, col)
            cell.value = None
            if cell.hyperlink:
                cell.hyperlink = None
            if cell.comment and row >= DATA_START_ROW and col != COL_MIGRATION_REQUEST_ID:
                # keep header comments only
                pass
    for row in range(DATA_START_ROW, end_row + 1):
        ws.cell(row, COL_MIGRATION_REQUEST_ID).value = migration_request_id

    if "Instructions" in wb.sheetnames:
        del wb["Instructions"]
    instructions = wb.create_sheet("Instructions", 0)
    instructions["A1"] = "IMPORTANT — migration_request_id"
    instructions["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    instructions["A1"].fill = PatternFill("solid", fgColor="B91C1C")
    instructions.merge_cells("A1:F1")
    instructions["A2"] = (
        f'This template is locked to migration_request_id = "{migration_request_id}". '
        "Do NOT change column A (migration_request_id) on the Scope sheet. "
        "When you upload the completed file, every filled data row must keep this exact ID "
        "or the upload will be rejected."
    )
    instructions["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions.merge_cells("A2:F4")
    instructions["A6"] = "How to fill L1 → L2 → L3 → L4"
    instructions["A6"].font = Font(bold=True, size=12)
    instructions["A7"] = (
        "1) Open the Scope sheet.\n"
        "2) Select L1 (column C) from the dropdown.\n"
        "3) Select L2 (column D) — options are filtered by L1.\n"
        "4) Select L3 (column E) — filtered by L1 + L2.\n"
        "5) Select L4 (column F) — filtered by L1 + L2 + L3.\n"
        "6) Fill the remaining task columns. Leave unused rows blank.\n"
        "7) Save the file and use Bulk Upload on the Opportunity Assessment page."
    )
    instructions["A7"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions.merge_cells("A7:F12")
    instructions.column_dimensions["A"].width = 28
    for letter in "BCDEF":
        instructions.column_dimensions[letter].width = 18

    l1_list, l2_by_l1, l3_by_l1_l2, l4_by_l1_l2_l3 = _load_cascade_maps()

    if "_CascadeLists" in wb.sheetnames:
        del wb["_CascadeLists"]
    lists = wb.create_sheet("_CascadeLists")

    for name in list(wb.defined_names):
        if name == "L1LIST" or name.startswith(("L2_", "L3_", "L4_")):
            del wb.defined_names[name]

    col = 1
    l1_ref = _write_list_column(lists, col, "L1", l1_list)
    _add_defined_name(wb, "L1LIST", l1_ref)
    col += 1

    for l1, l2_values in l2_by_l1.items():
        name = named_l2(l1)
        ref = _write_list_column(lists, col, name, l2_values)
        _add_defined_name(wb, name, ref)
        col += 1

    for (l1, l2), l3_values in l3_by_l1_l2.items():
        name = named_l3(l1, l2)
        ref = _write_list_column(lists, col, name, l3_values)
        _add_defined_name(wb, name, ref)
        col += 1

    for (l1, l2, l3), l4_values in l4_by_l1_l2_l3.items():
        name = named_l4(l1, l2, l3)
        ref = _write_list_column(lists, col, name, l4_values)
        _add_defined_name(wb, name, ref)
        col += 1

    lists.sheet_state = "hidden"

    ws.data_validations.dataValidation = []

    dv_l1 = DataValidation(
        type="list",
        formula1="=L1LIST",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid L1",
        error="Please select L1 from the dropdown list.",
    )
    dv_l1.add(f"C{DATA_START_ROW}:C{end_row}")
    ws.add_data_validation(dv_l1)

    dv_l2 = DataValidation(
        type="list",
        formula1=f'=INDIRECT("L2_"&{excel_sanitize_formula("C3")})',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid L2",
        error="Select L1 first, then choose L2 from the filtered list.",
    )
    dv_l2.add(f"D{DATA_START_ROW}:D{end_row}")
    ws.add_data_validation(dv_l2)

    dv_l3 = DataValidation(
        type="list",
        formula1=(
            f'=INDIRECT("L3_"&{excel_sanitize_formula("C3")}'
            f'&"_"&{excel_sanitize_formula("D3")})'
        ),
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid L3",
        error="Select L1 and L2 first, then choose L3.",
    )
    dv_l3.add(f"E{DATA_START_ROW}:E{end_row}")
    ws.add_data_validation(dv_l3)

    dv_l4 = DataValidation(
        type="list",
        formula1=(
            f'=INDIRECT("L4_"&{excel_sanitize_formula("C3")}'
            f'&"_"&{excel_sanitize_formula("D3")}'
            f'&"_"&{excel_sanitize_formula("E3")})'
        ),
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid L4",
        error="Select L1, L2 and L3 first, then choose L4.",
    )
    dv_l4.add(f"F{DATA_START_ROW}:F{end_row}")
    ws.add_data_validation(dv_l4)

    dv_complexity = DataValidation(
        type="list",
        formula1='"Low,Medium,High"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_complexity.add(f"N{DATA_START_ROW}:N{end_row}")
    ws.add_data_validation(dv_complexity)

    ws.freeze_panes = "A3"
    ws["A2"].comment = Comment(
        f'Do not change migration_request_id. Must remain: "{migration_request_id}"',
        "WPM",
        width=320,
        height=70,
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def parse_opportunity_workbook(file_obj, expected_migration_request_id: str) -> dict:
    wb = load_workbook(file_obj, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name in {"Instructions", "_CascadeLists"}:
            continue
        if name == "Scope":
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            if name not in {"Instructions", "_CascadeLists"}:
                ws = wb[name]
                break
    if ws is None:
        raise ValueError("No data sheet found in workbook.")

    headers = [_normalize_header(ws.cell(HEADER_ROW, c).value) for c in range(1, 40)]
    col_map: dict[str, int] = {}
    for idx, header in enumerate(headers, start=1):
        field = FIELD_BY_HEADER.get(header)
        if field:
            col_map[field] = idx

    if "migration_request_id" not in col_map:
        raise ValueError("Missing migration_request_id column in header row.")

    expected = (expected_migration_request_id or "").strip()
    rows: list[dict] = []
    errors: list[dict] = []

    def cell(row_idx: int, field: str) -> str:
        col = col_map.get(field)
        if not col:
            return ""
        value = ws.cell(row_idx, col).value
        if value is None:
            return ""
        return str(value).strip()

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        payload = {
            "migration_request_id": cell(row_idx, "migration_request_id"),
            "owner": cell(row_idx, "owner"),
            "product": cell(row_idx, "product"),
            "location": cell(row_idx, "location"),
            "l1": cell(row_idx, "l1"),
            "l2": cell(row_idx, "l2"),
            "l3": cell(row_idx, "l3"),
            "l4": cell(row_idx, "l4"),
            "task_name": cell(row_idx, "task_name"),
            "task_description": cell(row_idx, "task_description"),
            "upstream": cell(row_idx, "upstream"),
            "downstream": cell(row_idx, "downstream"),
            "risks_related": cell(row_idx, "risks_related"),
            "complexity": cell(row_idx, "complexity"),
            "sop_iop_exists": cell(row_idx, "sop_iop_exists"),
            "training_time_needed": cell(row_idx, "training_time_needed"),
            "recommended_handoff_duration": cell(row_idx, "recommended_handoff_duration"),
            "task_frequency": cell(row_idx, "task_frequency"),
            "unit_of_measure": cell(row_idx, "unit_of_measure"),
            "volume_monthly": cell(row_idx, "volume_monthly"),
            "task_time_per_unit_min": cell(row_idx, "task_time_per_unit_min"),
            "task_found_in_service_catalog": cell(row_idx, "task_found_in_service_catalog"),
            "migratable_to_gsc": cell(row_idx, "migratable_to_gsc"),
            "fte_calculation": cell(row_idx, "fte_calculation"),
        }

        content_fields = [
            "product",
            "owner",
            "location",
            "l1",
            "l2",
            "l3",
            "l4",
            "task_name",
            "task_description",
            "upstream",
            "downstream",
            "risks_related",
            "complexity",
            "training_time_needed",
            "recommended_handoff_duration",
            "task_frequency",
            "unit_of_measure",
            "volume_monthly",
            "task_time_per_unit_min",
            "task_found_in_service_catalog",
            "migratable_to_gsc",
            "fte_calculation",
        ]
        # Ignore rows that only have migration_request_id (or leftover noise)
        if not any(payload[k] for k in content_fields):
            continue

        mid = payload["migration_request_id"]
        if mid != expected:
            errors.append(
                {
                    "row": row_idx,
                    "error": (
                        f"migration_request_id mismatch: got '{mid or '(empty)'}', "
                        f"expected '{expected}'."
                    ),
                }
            )
            continue

        rows.append(payload)

    return {
        "rows": rows,
        "errors": errors,
        "accepted_count": len(rows),
        "rejected_count": len(errors),
    }
